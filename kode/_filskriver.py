import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

fill1 = PatternFill(fill_type='solid',start_color='8AB8EA')
fill2 = PatternFill(fill_type='solid',start_color='DADCDC')
fill3 = PatternFill(fill_type='solid',start_color='041465')

font1 = Font(name='Arial',size=10,bold=True,color='FFFFFF')
font2 = Font(name='Arial',size=10,bold=True,color='974706')
font3 = Font(name='Arial',size=18,bold=True,color='974706')
font4 = Font(name='Arial',size=12,bold=True,color='000000')
font5 = Font(name='Arial',size=14,bold=True,color='FFFFFF')

thick_around = Border(left=Side(border_style='thick'),right=Side(border_style='thick'),top=Side(border_style='thick'),bottom=Side(border_style='thick'))
thick_left_else_thin = Border(left=Side(border_style='thick'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
thick_right_else_thin = Border(left=Side(border_style='thin'),right=Side(border_style='thick'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
thick_top_else_thin = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thick'),bottom=Side(border_style='thin'))
thick_top = Border(top=Side(border_style='thick'))
thin_around = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
thick_left_top_corner = Border(left=Side(border_style='thick'),right=Side(border_style='thin'),top=Side(border_style='thick'),bottom=Side(border_style='thin'))
thick_left_down_corner = Border(left=Side(border_style='thick'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thick'))
thick_right_top_corner = Border(left=Side(border_style='thin'),right=Side(border_style='thick'),top=Side(border_style='thick'),bottom=Side(border_style='thin'))
thick_right_down_corner = Border(left=Side(border_style='thin'),right=Side(border_style='thick'),top=Side(border_style='thin'),bottom=Side(border_style='thick'))


class Filskriver:

    @staticmethod
    def tabellhistorie(datasenter):

        innstillinger = datasenter.innstillinger()
        oppdater = innstillinger["oppdater tabellhistorie"]

        if not oppdater:
            return

        aar = datasenter.aar()

        with open(f"./data/tabellhistorie_{aar}.csv","a") as fil:

            fil.write("|" + str(datasenter.inititieringTid()) + "| | |\n")

            for kjonn in ["menn","kvinner"]:
                tabell = datasenter.tabell(kjonn)
                for div in [1,2,3]:
                    for lag in tabell.hentDivisjon(div)["Tabell"]:

                        fil.write(kjonn + "|")
                        fil.write(str(div-1) + "|")
                        fil.write(str(lag.hentPosisjon()) + "|")
                        fil.write(str(lag.hentPoeng()) + "|")
                        fil.write(lag.hentLagnavn() + "\n")

            fil.close()


    @staticmethod
    def resultatark(datasenter,filnavn):

        dokument = openpyxl.Workbook()
        del dokument['Sheet']

        ark = {"menn": {}, "kvinner": {}}

        ark["menn"]["stat"] = dokument.create_sheet(title='Menn - statistikk', index=0)
        ark["menn"]["avvik"] = dokument.create_sheet(title='resultatavvik', index=1)
        ark["menn"]["klubbres"] = dokument.create_sheet(title='klubbres', index=2)

        ark["kvinner"]["stat"] = dokument.create_sheet(title='Kvinner - statistikk', index=3)
        ark["kvinner"]["avvik"] = dokument.create_sheet(title='resuItatavvik', index=4)
        ark["kvinner"]["klubbres"] = dokument.create_sheet(title='kIubbres', index=5)

        for kjonn in ["menn","kvinner"]:
            
            # statistikk

            statark = ark[kjonn]["stat"]
            statark.cell(row=2,column=2, value="Statistikk "+kjonn)

            for i,kolonnetittel in enumerate(["Navn","F. år","Øvelse","Resultat","Poeng","Klubb","Krets","Sted","Dato"]):
                statark.cell(row=4,column=2+i, value=kolonnetittel)

            i = 0
            for res in datasenter.resultater(kjonn):

                if (klubbFra := res.hentKlubbFra()) == None:
                    continue

                statark.cell(row=5+i,column=2, value=res.hentUtover().hentNavn())
                statark.cell(row=5+i,column=3, value=res.hentUtover().hentFAar())
                statark.cell(row=5+i,column=4, value=res.hentOvelse())
                statark.cell(row=5+i,column=5, value=res.hentPrestasjon())
                statark.cell(row=5+i,column=6, value=res.hentPoeng())
                statark.cell(row=5+i,column=7, value=klubbFra.hentKlubbnavn())
                statark.cell(row=5+i,column=8, value=klubbFra.hentKrets())
                statark.cell(row=5+i,column=9, value=res.hentSted())
                statark.cell(row=5+i,column=10, value=res.hentDato())

                i += 1

            # resultatavvik

            avvikark = ark[kjonn]["avvik"]
            avvikark.cell(row=2,column=2, value="Resultatavvik "+kjonn)
            avvik = datasenter.resultater(kjonn).avvik()

            for i,kolonnetittel in enumerate(["Fra klubb","Til klubb","Navn","F. år","Øvelse","Resultat","Poeng","Sted","Dato","Begrunnelse"]):
                avvikark.cell(row=4,column=2+i, value=kolonnetittel)

            for i,res in enumerate(avvik):

                if (klubbFra := res.hentKlubbFra()) != None:
                    avvikark.cell(row=5+i,column=2, value=klubbFra.hentKlubbnavn())
                else:
                    avvikark.cell(row=5+i,column=2, value="-")
                
                if (klubbTil := res.hentKlubbTil()) != None:
                    avvikark.cell(row=5+i,column=3, value=klubbTil.hentKlubbnavn())
                else:
                    avvikark.cell(row=5+i,column=3, value="-")

                avvikark.cell(row=5+i,column=4, value=res.hentUtover().hentNavn())
                avvikark.cell(row=5+i,column=5, value=res.hentUtover().hentFAar())
                avvikark.cell(row=5+i,column=6, value=res.hentOvelse())
                avvikark.cell(row=5+i,column=7, value=res.hentPrestasjon())
                avvikark.cell(row=5+i,column=8, value=res.hentPoeng())
                avvikark.cell(row=5+i,column=9, value=res.hentSted())
                avvikark.cell(row=5+i,column=10, value=res.hentDato())
                avvikark.cell(row=5+i,column=11, value=res.hentBegrunnelse())

            # klubbresultater

            rad = 2

            klubbark = ark[kjonn]["klubbres"]
            klubbark.cell(row=rad,column=2, value="Klubbresultater "+kjonn)
            klubber = datasenter.klubber(kjonn)

            for klubb in klubber:

                if not klubb.harResultater():
                    continue

                rad += 3

                klubbark.cell(row=rad,column=2, value="Klubb")
                klubbark.cell(row=rad,column=3, value=klubb.hentKlubbnavn())
                klubbark.cell(row=rad+1,column=2, value="Krets")
                klubbark.cell(row=rad+1,column=3, value=klubb.hentKrets())

                rad += 3

                for i,kolonnetittel in enumerate(["Navn","F. år","Øvelse","Resultat","Poeng","Sted","Dato"]):
                    klubbark.cell(row=rad,column=2+i, value=kolonnetittel)

                for res in klubb.hentResultater().set():
                    rad += 1

                    klubbark.cell(row=rad,column=2, value=res.hentUtover().hentNavn())
                    klubbark.cell(row=rad,column=3, value=res.hentUtover().hentFAar())
                    klubbark.cell(row=rad,column=4, value=res.hentOvelse())
                    klubbark.cell(row=rad,column=5, value=res.hentPrestasjon())
                    klubbark.cell(row=rad,column=6, value=res.hentPoeng())
                    klubbark.cell(row=rad,column=7, value=res.hentSted())
                    klubbark.cell(row=rad,column=8, value=res.hentDato())


        dokument.save(f"./output/{filnavn}.xlsx")

    @staticmethod
    def _formaterCelle(ark,rad,kol,verdi=None,font=None,fyll=None,kant=None):
    
        if verdi != None:
            ark.cell(row=rad,column=kol, value=verdi)
        if font != None:
            ark.cell(row=rad,column=kol).font = font
        if fyll != None:
            ark.cell(row=rad,column=kol).fill = fyll
        if kant != None:
            ark.cell(row=rad,column=kol).border = kant 

    @staticmethod
    def offisieltSerieark(datasenter,filnavn):

        aar = datasenter.aar()        

        "Definerer filer og ark i egen dict"

        """
        col = ws.column_dimensions['A']
        col.font = Font(bold=True)
        row = ws.row_dimensions[1]
        row.font = Font(underline="single")
        """

        div_1_2 = openpyxl.Workbook()
        menn_3 = openpyxl.Workbook()
        kvinner_3 = openpyxl.Workbook()

        del div_1_2['Sheet']
        del menn_3['Sheet']
        del kvinner_3['Sheet']
        
        kretser = datasenter.hentAlleKretser()
        
        ark = {"menn": {"1.d": {}, "2.d": {}, "3.d": {}},"kvinner": {"1.d": {}, "2.d": {}, "3.d": {}}}

        # lager faner

        nr = 0
        for kj,dokument in zip(["menn","kvinner"],[menn_3,kvinner_3]):

            # 3. divisjon

            ark[kj]["3.d"]["Tabell"] = dokument.create_sheet(title="Tabell", index=0)
            for i,krets in enumerate(kretser):
                ark[kj]["3.d"][krets] = dokument.create_sheet(title=krets, index=1+i)

            # 1. og 2. divisjon

            for div in ["1.d","2.d"]:
                for fane in ["Tabell","Detaljer"]:
                    ark[kj][div][fane] = div_1_2.create_sheet(title=f"{fane} {kj} {div}", index=nr)
                    nr += 1

        for kj in ["menn","kvinner"]:

            iKrets = {}
            for krets in kretser:
                iKrets[krets] = 0

            "Fyller alle excelark med verdier"

            tabell = datasenter.tabell(kj)

            for div in [1,2,3]:

                settinger = datasenter.settinger()

                N_obl = settinger["antall øvelser"][f"{div}. div"]["obl"]
                N_val = settinger["antall øvelser"][f"{div}. div"]["val"]

                # setter opp tabellark

                divisjon = tabell.hentDivisjon(div)["Tabell"]
                tabellark = ark[kj][f"{div}.d"]["Tabell"]
                
                tabellark.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
                tabellark.cell(row=1, column=1).fill = fill3
                tabellark.cell(row=1, column=1,value=str(div)+'. DIVISJON ' + kj.upper() + ' ' + aar + ' TOTALT')
                tabellark.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                tabellark.cell(row=1, column=1).font = font5  
                
                tabellark.row_dimensions[1].height = 18
                for rad in range(2,len(divisjon)+3):
                    tabellark.row_dimensions[rad].height = 13
                
                for kol,bredde in enumerate([4.36,24.55,7.64,8.73,7.64,16.45]):
                    tabellark.column_dimensions[chr(65+kol)].width = bredde

                for kol in range(26):
                    kol = tabellark.column_dimensions[chr(65+kol)]
                    kol.fill = fill1

                for arknavn,detaljark in ark[kj][f"{div}.d"].items():
                    if arknavn=="Tabell":
                        continue

                    for kol in range(26):
                        kol = detaljark.column_dimensions[chr(65+kol)]
                        kol.fill = fill1

                    for side in [0,1]:
                        for kol,bredde in enumerate([0.56,19.36,38.82,7.73,10.73,9.18,21.27,6.36,0.56,4.73]):
                            kolonne = chr(65+10*side+kol)
                            detaljark.column_dimensions[kolonne].width = bredde        

                # setter opp lagoppsett-ark
            
                for lag in divisjon:
                    
                    " Setter inn overordnet laginfo i tilhørende tabellark "
                    
                    div = lag.hentDiv()
                    tabellark = ark[kj][f"{div}.d"]["Tabell"]
                    krets = lag.hentKrets()

                    pos = lag.hentPosisjon()
                    
                    Filskriver._formaterCelle(tabellark,pos+3,1, font=font2, fyll=fill2, verdi=pos+1)
                    Filskriver._formaterCelle(tabellark,pos+3,2, font=font2, fyll=fill2, verdi=lag.hentLagnavn())
                    Filskriver._formaterCelle(tabellark,pos+3,3, font=font2, fyll=fill2, verdi='('+lag.hentNotat()+')')
                    Filskriver._formaterCelle(tabellark,pos+3,4, font=font2, fyll=fill2, verdi=lag.hentPoeng())
                    Filskriver._formaterCelle(tabellark,pos+3,5, font=font2, fyll=fill2, verdi="("+lag.hentFjoraarsplassering()+")")
                    Filskriver._formaterCelle(tabellark,pos+3,6, font=font2, fyll=fill2, verdi=krets)
                    
                    tabellark.cell(row=pos+3, column=5).alignment = Alignment(horizontal='right')

                    """
                    if not laginfo["f. plassering"][-1] in ["d","-"]:
                        tabell.cell(row=i+3, column=5,value=int(laginfo["f. plassering"]))
                        tabell.cell(row=i+3, column=5).number_format = '##('+len(str(laginfo["f. plassering"]))*'0'+")"
                    else:
                        tabell.cell(row=i+3, column=5,value="("+laginfo["f. plassering"]+")")
                    tabell.cell(row=i+3, column=5).alignment = Alignment(horizontal='right')
                    """
                    
                    " Setter inn detaljert laginfo i tilhørende detaljark "
                    
                    if div==3:
                        pos = iKrets[krets]
                        iKrets[krets] += 1

                        detaljark = ark[kj]["3.d"][krets]
                    else:
                        detaljark = ark[kj][f"{div}.d"]["Detaljer"]
                    
                    al = pos % 2
                    
                    rad = int((1/2)*(pos-al))*(N_obl+N_val+22)+1
                    kol = 10*al
                        
                    for h in range(rad,rad+N_obl+N_val+21):
                        detaljark.cell(row=h, column=kol+10).fill = fill3
                        detaljark.row_dimensions[h].height = 13
                    detaljark.row_dimensions[rad+N_obl+N_val+21].height = 13
                    for h in range(1,11):
                        detaljark.cell(row=rad+N_obl+N_val+21, column=kol+h).fill = fill3      
                            
                    " Lager obligatorisk og valgfri resultatbokser "
                        
                    for k in [0,N_obl+6]:
                        for p in [0,1,2,3,4,5,6]:
                            detaljark.cell(row=rad+k+7, column=kol+p+2).fill = fill3
                            detaljark.cell(row=rad+k+7, column=kol+p+2).font = font1
                        for p in [1,2,3,4,5]:
                            detaljark.cell(row=rad+k+7, column=kol+p+2).border = thick_top_else_thin
                                
                        detaljark.cell(row=rad+k+7, column=kol+2).border = thick_left_top_corner
                        detaljark.cell(row=rad+k+7, column=kol+8).border = thick_right_top_corner
                            
                        detaljark.cell(row=rad+k+7, column=kol+2,value='ØVELSE')
                        detaljark.cell(row=rad+k+7, column=kol+3,value='NAVN')               
                        detaljark.cell(row=rad+k+7, column=kol+4,value='FØDT ÅR')
                        detaljark.cell(row=rad+k+7, column=kol+5,value='RESULTAT')
                        detaljark.cell(row=rad+k+7, column=kol+6,value='POENG')
                        detaljark.cell(row=rad+k+7, column=kol+7,value='STED')
                        detaljark.cell(row=rad+k+7, column=kol+8,value='DATO')
                            
                    Filskriver._formaterCelle(detaljark,rad+1,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="Krets")
                    Filskriver._formaterCelle(detaljark,rad+2,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="Klubb")
                    Filskriver._formaterCelle(detaljark,rad+3,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="År")
                    
                    Filskriver._formaterCelle(detaljark,rad+1,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=krets)
                    Filskriver._formaterCelle(detaljark,rad+2,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=lag.hentLagnavn())
                    Filskriver._formaterCelle(detaljark,rad+3,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=int(aar))
                    detaljark.cell(row=rad+3, column=kol+3).alignment = Alignment(horizontal='left')

                    Filskriver._formaterCelle(detaljark,rad+5,kol+2, font=font4, fyll=fill1, verdi='OBLIGATORISKE ØVELSER')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+11,kol+2, font=font4, fyll=fill1, verdi='VALGFRIE ØVELSER')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+2, font=font1, fyll=fill3, verdi='Antall noteringer:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+5, font=font1, fyll=fill3, verdi='Sum:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="obl")["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="obl")["poeng"])
            
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+2, font=font1, fyll=fill3, verdi='Antall noteringer:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+5, font=font1, fyll=fill3, verdi='Sum:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="val")["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="val")["poeng"])
            
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+2, font=font1, fyll=fill3, verdi='Noteringer totalt:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+5, font=font1, fyll=fill3, verdi='Totalsum:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+19,kol+2, font=font1, fyll=fill3, verdi='Antall deltakere:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["poeng"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+19,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["antall deltakere"])

                    for k,type in zip([0,N_obl+6],["obl","val"]):

                        oppstilling = lag.hentOppstilling(type=type)["resultater"]

                        for oi,res in enumerate(oppstilling):                 
                            
                            detaljark.cell(row=rad+k+oi+8, column=kol+2,value=res.hentOvelse())
                            detaljark.cell(row=rad+k+oi+8, column=kol+3,value=res.hentUtover().hentNavn())

                            detaljark.cell(row=rad+k+oi+8, column=kol+4,value=int(res.hentUtover().hentFAar()))

                            try:
                                resultat = res.hentPrestasjon().replace(",",".")
                                detaljark.cell(row=rad+k+oi+8, column=kol+5,value=float(resultat))
                                if len(resultat.split(".")[-1])==2:
                                    detaljark.cell(row=rad+k+oi+8, column=kol+5).number_format = '#,##0.00'
                                else:
                                    detaljark.cell(row=rad+k+oi+8, column=kol+5).number_format = '#,#0.0'
                            except ValueError:
                                detaljark.cell(row=rad+k+oi+8, column=kol+5,value=str(res.hentPrestasjon()))
                                detaljark.cell(row=rad+k+oi+8, column=kol+5).alignment = Alignment(horizontal='right')

                            detaljark.cell(row=rad+k+oi+8, column=kol+6,value=res.hentPoeng())
                            detaljark.cell(row=rad+k+oi+8, column=kol+7,value=res.hentSted())
                                    
                            
                            try:
                                resultat = res.hentDato()[:-5].replace(",",".")
                                detaljark.cell(row=rad+k+oi+8, column=kol+8,value=float(resultat))
                                detaljark.cell(row=rad+k+oi+8, column=kol+8).number_format = '##,##00.00'
                                detaljark.cell(row=rad+k+oi+8, column=kol+8).alignment = Alignment(horizontal='left')
                            except ValueError:
                                detaljark.cell(row=rad+k+oi+8, column=kol+8,value=res.hentDato()[:-5])
                            
                        
                        
                    for k,l in zip([0,N_obl+6],[N_obl,N_val]):
                        for oi in range(l):
                            Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+2, font=font2, fyll=fill2, kant=thick_left_else_thin)
                            Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+8, font=font2, fyll=fill2, kant=thick_right_else_thin)
                            for h in [1,2,3,4,5]:
                                Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+h+2, font=font2, fyll=fill2, kant=thin_around)

                        for oi in range(2,9):
                            detaljark.cell(row=rad+l+k+8, column=kol+oi).border = thick_top
                            detaljark.cell(row=rad+l+k+8, column=kol+oi).fill = fill1
                    detaljark.merge_cells(start_row=rad+1, start_column=kol+6, end_row=rad+2, end_column=kol+7)
                        
                    Filskriver._formaterCelle(detaljark,rad+1,kol+6, font=font3, fyll=fill2, verdi=lag.hentPosisjon()+1)
                    detaljark.cell(row=rad+1, column=kol+6).alignment = Alignment(horizontal='center')
                
        div_1_2.save(f"./output/{filnavn} 1-2. div.xlsx")
        menn_3.save(f"./output/{filnavn} menn 3. div.xlsx")
        kvinner_3.save(f"./output/{filnavn} kvinner 3. div.xlsx")

    @staticmethod
    def utviklingSerietabell(datasenter,filnavn):

        aar = datasenter.aar()        

        "Definerer filer og ark i egen dict"

        div_1_2 = openpyxl.Workbook()
        menn_3 = openpyxl.Workbook()
        kvinner_3 = openpyxl.Workbook()

        del div_1_2['Sheet']
        del menn_3['Sheet']
        del kvinner_3['Sheet']
        
        kretser = datasenter.hentAlleKretser()
        
        ark = {"menn": {"1.d": {}, "2.d": {}, "3.d": {}},"kvinner": {"1.d": {}, "2.d": {}, "3.d": {}}}

        # lager faner

        nr = 0
        for kj,dokument in zip(["menn","kvinner"],[menn_3,kvinner_3]):

            # 3. divisjon

            ark[kj]["3.d"]["Tabell"] = dokument.create_sheet(title="Tabell", index=0)
            for i,krets in enumerate(kretser):
                ark[kj]["3.d"][krets] = dokument.create_sheet(title=krets, index=1+i)

            # 1. og 2. divisjon

            for div in ["1.d","2.d"]:
                for fane in ["Tabell","Detaljer"]:
                    ark[kj][div][fane] = div_1_2.create_sheet(title=f"{fane} {kj} {div}", index=nr)
                    nr += 1

            
        for kj in ["menn","kvinner"]:

            iKrets = {}
            for krets in kretser:
                iKrets[krets] = 0

            "Fyller alle excelark med verdier"

            tabell = datasenter.tabell(kj)

            for div in [1,2,3]:

                settinger = datasenter.settinger()

                N_obl = settinger["antall øvelser"][f"{div}. div"]["obl"]
                N_val = settinger["antall øvelser"][f"{div}. div"]["val"]

                # setter opp tabellark

                divisjon = tabell.hentDivisjon(div)["Tabell"]
                tabellark = ark[kj][f"{div}.d"]["Tabell"]
                
                tabellark.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
                tabellark.cell(row=1, column=1).fill = fill3
                tabellark.cell(row=1, column=1,value=str(div)+'. DIVISJON MENN ' + aar + ' TOTALT')
                tabellark.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                tabellark.cell(row=1, column=1).font = font5  

                # setter opp lagoppsett-ark
            
                for lag in divisjon:
                    
                    " Setter inn overordnet laginfo i tilhørende tabellark "
                    
                    div = lag.hentDiv()

                    tabellark = ark[kj][f"{div}.d"]["Tabell"]
                    krets = lag.hentKrets()

                    pos = lag.hentPosisjon()
                    
                    Filskriver._formaterCelle(tabellark,pos+3,1, font=font2, fyll=fill2, verdi=pos+1)
                    Filskriver._formaterCelle(tabellark,pos+3,2, font=font2, fyll=fill2, verdi=lag.hentLagnavn())
                    Filskriver._formaterCelle(tabellark,pos+3,3, font=font2, fyll=fill2, verdi='('+lag.hentNotat()+')')
                    Filskriver._formaterCelle(tabellark,pos+3,4, font=font2, fyll=fill2, verdi=lag.hentPoeng())
                    Filskriver._formaterCelle(tabellark,pos+3,5, font=font2, fyll=fill2)
                    Filskriver._formaterCelle(tabellark,pos+3,6, font=font2, fyll=fill2, verdi=krets)
                    
                    """
                    if not laginfo["f. plassering"][-1] in ["d","-"]:
                        tabell.cell(row=i+3, column=5,value=int(laginfo["f. plassering"]))
                        tabell.cell(row=i+3, column=5).number_format = '##('+len(str(laginfo["f. plassering"]))*'0'+")"
                    else:
                        tabell.cell(row=i+3, column=5,value="("+laginfo["f. plassering"]+")")
                    tabell.cell(row=i+3, column=5).alignment = Alignment(horizontal='right')
                    """
                    
                    " Setter inn detaljert laginfo i tilhørende detaljark "
                    
                    if div==3:
                        pos = iKrets[krets]
                        iKrets[krets] += 1

                        detaljark = ark[kj]["3.d"][krets]
                    else:
                        detaljark = ark[kj][f"{div}.d"]["Detaljer"]
                    
                    al = pos % 2
                    
                    rad = int((1/2)*(pos-al))*(N_obl+N_val+22)+1
                    kol = 10*al
                        
                    for h in range(rad,rad+N_obl+N_val+21):
                        detaljark.cell(row=h, column=kol+10).fill = fill3
                        detaljark.row_dimensions[h].height = 13
                    detaljark.row_dimensions[rad+N_obl+N_val+21].height = 13
                    for h in range(1,11):
                        detaljark.cell(row=rad+N_obl+N_val+21, column=kol+h).fill = fill3      
                            
                    " Lager obligatorisk og valgfri resultatbokser "
                        
                    for k in [0,N_obl+6]:
                        for p in [0,1,2,3,4,5,6]:
                            detaljark.cell(row=rad+k+7, column=kol+p+2).fill = fill3
                            detaljark.cell(row=rad+k+7, column=kol+p+2).font = font1
                        for p in [1,2,3,4,5]:
                            detaljark.cell(row=rad+k+7, column=kol+p+2).border = thick_top_else_thin
                                
                        detaljark.cell(row=rad+k+7, column=kol+2).border = thick_left_top_corner
                        detaljark.cell(row=rad+k+7, column=kol+8).border = thick_right_top_corner
                            
                        detaljark.cell(row=rad+k+7, column=kol+2,value='ØVELSE')
                        detaljark.cell(row=rad+k+7, column=kol+3,value='NAVN')               
                        detaljark.cell(row=rad+k+7, column=kol+4,value='FØDT ÅR')
                        detaljark.cell(row=rad+k+7, column=kol+5,value='RESULTAT')
                        detaljark.cell(row=rad+k+7, column=kol+6,value='POENG')
                        detaljark.cell(row=rad+k+7, column=kol+7,value='STED')
                        detaljark.cell(row=rad+k+7, column=kol+8,value='DATO')
                            
                    Filskriver._formaterCelle(detaljark,rad+1,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="Krets")
                    Filskriver._formaterCelle(detaljark,rad+2,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="Klubb")
                    Filskriver._formaterCelle(detaljark,rad+3,kol+2, font=font1, fyll=fill3, kant=thin_around, verdi="År")
                    
                    Filskriver._formaterCelle(detaljark,rad+1,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=krets)
                    Filskriver._formaterCelle(detaljark,rad+2,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=lag.hentLagnavn())
                    Filskriver._formaterCelle(detaljark,rad+3,kol+3, font=font2, fyll=fill2, kant=thin_around, verdi=aar)
                    detaljark.cell(row=rad+3, column=kol+3).alignment = Alignment(horizontal='left')

                    Filskriver._formaterCelle(detaljark,rad+5,kol+2, font=font4, fyll=fill1, verdi='OBLIGATORISKE ØVELSER')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+11,kol+2, font=font4, fyll=fill1, verdi='VALGFRIE ØVELSER')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+2, font=font1, fyll=fill3, verdi='Antall noteringer:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+5, font=font1, fyll=fill3, verdi='Sum:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="obl")["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+9,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="obl")["poeng"])
            
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+2, font=font1, fyll=fill3, verdi='Antall noteringer:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+5, font=font1, fyll=fill3, verdi='Sum:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="val")["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+15,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling(type="val")["poeng"])
            
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+2, font=font1, fyll=fill3, verdi='Noteringer totalt:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+5, font=font1, fyll=fill3, verdi='Totalsum:')
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+19,kol+2, font=font1, fyll=fill3, verdi='Antall deltakere:')
                        
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["antall resultater"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+17,kol+6, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["poeng"])
                    Filskriver._formaterCelle(detaljark,rad+N_obl+N_val+19,kol+3, font=font2, fyll=fill2, kant=thick_around, verdi=lag.hentOppstilling()["antall deltakere"])

                    for k,type in zip([0,N_obl+6],["obl","val"]):

                        oppstilling = lag.hentOppstilling(type=type)["resultater"]

                        for oi,res in enumerate(oppstilling):                    
                            
                            detaljark.cell(row=rad+k+oi+8, column=kol+2,value=res.hentOvelse())
                            detaljark.cell(row=rad+k+oi+8, column=kol+3,value=res.hentUtover().hentNavn())

                            """
                            try:
                                detaljark.cell(row=rad+k+oi+8, column=kol+4,value=int(res["f. år"]))                
                            except:
                                detaljark.cell(row=rad+k+oi+8, column=kol+4,value=str(res["f. år"]))
                                detaljark.cell(row=rad+k+oi+8, column=kol+4).alignment = Alignment(horizontal='right')
                            try:
                                resultat = res["resultat"].replace(",",".")
                                detaljark.cell(row=rad+k+oi+8, column=kol+5,value=float(resultat))
                                if len(resultat.split(".")[-1])==2:
                                    detaljark.cell(row=rad+k+oi+8, column=kol+5).number_format = '#,##0.00'
                                else:
                                    detaljark.cell(row=rad+k+oi+8, column=kol+5).number_format = '#,#0.0'
                            except:
                                detaljark.cell(row=rad+k+oi+8, column=kol+5,value=str(res["resultat"]))
                                detaljark.cell(row=rad+k+oi+8, column=kol+5).alignment = Alignment(horizontal='right')
                            """

                            detaljark.cell(row=rad+k+oi+8, column=kol+6,value=res.hentPoeng())
                            detaljark.cell(row=rad+k+oi+8, column=kol+7,value=res.hentSted())
                                    
                            """
                            try:
                                resultat = res["dato"][:-5].replace(",",".")
                                detaljark.cell(row=rad+k+oi+8, column=kol+8,value=float(resultat))
                                detaljark.cell(row=rad+k+oi+8, column=kol+8).number_format = '##,##00.00'
                                detaljark.cell(row=rad+k+oi+8, column=kol+8).alignment = Alignment(horizontal='left')
                            except:
                                detaljark.cell(row=rad+k+oi+8, column=kol+8,value=res["dato"][:-5])
                            """
                        
                        
                    for k,l in zip([0,N_obl+6],[N_obl,N_val]):
                        for oi in range(l):
                            Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+2, font=font2, fyll=fill2, kant=thick_left_else_thin)
                            Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+8, font=font2, fyll=fill2, kant=thick_right_else_thin)
                            for h in [1,2,3,4,5]:
                                Filskriver._formaterCelle(detaljark,rad+k+oi+8,kol+h+2, font=font2, fyll=fill2, kant=thin_around)

                        for oi in range(2,9):
                            detaljark.cell(row=rad+N_obl+k+8, column=kol+oi).border = thick_top
                            detaljark.cell(row=rad+N_obl+k+8, column=kol+oi).fill = fill1
                    detaljark.merge_cells(start_row=rad+1, start_column=kol+6, end_row=rad+2, end_column=kol+7)
                        
                    Filskriver._formaterCelle(detaljark,rad+1,kol+6, font=font3, fyll=fill2, verdi=lag.hentPosisjon()+1)
                    detaljark.cell(row=rad+1, column=kol+6).alignment = Alignment(horizontal='center')
                
        div_1_2.save(f"./output/{filnavn} 1-2. div.xlsx")
        menn_3.save(f"./output/{filnavn} menn 3. div.xlsx")
        kvinner_3.save(f"./output/{filnavn} kvinner 3. div.xlsx")
    