from _resultat import Resultat

from bs4 import BeautifulSoup
import requests
import openpyxl

url = "http://www.minfriidrettsstatistikk.info/php/SeriePoengPrKlubb.php"

class Statistikkhenting:

    @staticmethod
    def definerKlubbIDer(datasenter):
                
        try:
            data = BeautifulSoup(requests.get(url, timeout=10).text, "lxml")
        except (requests.ConnectionError, requests.Timeout):
            print(f"Mangler nett-tilgang og får ikke hentet klubbid-er.")
            return      
        klubber = data.find("select",{"name":"showclub"})
        
        unnga = True
        for html_linje in klubber:
            
            if unnga:
                unnga = False
                continue

            ID = html_linje["value"]
            klubbnavn = str(html_linje).replace('<option value="'+ID+'">','').replace('</option>','')

            klubbnavn = klubbnavn.replace("&amp;","&")
            
            for kjonn in ["menn","kvinner"]:
            
                klubb = datasenter.hentKlubbFraNavn(kjonn,klubbnavn)
                klubb.settID(ID)

    @staticmethod
    def hentKlubbstatistikk(datasenter,klubb):

        ovelsesinfo = datasenter.ovelsesinfo()
        
        ovelser_statistikk = ovelsesinfo["i statistikk"]
        ovelser_sluttform = ovelsesinfo["sluttform"]
        
        ID = klubb.hentID()
        kjonn = klubb.hentKjonn()
        aar = datasenter.aar()
        
        gender = "M" if kjonn=="menn" else "W" if kjonn=="kvinner" else None
            
        inputs = {'showclub': ID, 'showgender': gender, 'showyear': str(aar), "submit": "BEREGN"}

        while True:

            try:
                data = BeautifulSoup(requests.post(url, data=inputs, timeout=10).text, "lxml")
            except (requests.ConnectionError, requests.Timeout):
                print(f"Mangler nett-tilgang og får ikke hentet statistikk til {klubb}.")
                return                       
            all_data = data.findAll("table",{"id":"liten"})
            
            try:
                samlet_indiv_data = all_data[1].findAll("tr")[1:]
                break
            except IndexError:
                continue
            
        lag_data = [el.text for el in all_data[0].findAll("td")]
        
        klubb.settKrets(lag_data[1].replace(" Friidrettskrets",""))  

   
        data = [[el.text for el in indiv_data.findAll("td")] for indiv_data in samlet_indiv_data]            
            
        for ovelse,navn,fAar,res,poeng,sted,dato in data:
            
            skip = False
            for i in range(len(ovelser_statistikk)):
                for ov_st in ovelser_statistikk[i]:
                    if ov_st == ovelse:
                        ovelse = list(ovelser_sluttform.keys())[i]
                        skip = True
                        break
                if skip:
                    break
            
            utover = datasenter.hentUtover(kjonn,navn,fAar)
            
            resultat = Resultat(utover,ovelse,poeng,res,dato,sted,klubb)

            utover.leggTilRes(resultat)
            klubb.leggTilRes(resultat)
            datasenter.leggTilRes(kjonn,resultat)       

    @staticmethod
    def hentStatistikkFraFil(datasenter,filnavn):

        ovelsesinfo = datasenter.ovelsesinfo()["sluttform"]

        statistikkfil = openpyxl.load_workbook(f"./input/{filnavn}.xlsx")

        for kjonn in ["menn","kvinner"]:
            ark = statistikkfil[kjonn.capitalize() + " - statistikk"]
            for rad in range(5, ark.max_row+1):

                if rad%2500==0:
                    print(kjonn,rad)

                if ark.cell(row=rad,column=2).value == None:
                    continue

                klubbnavn = str(ark.cell(row=rad,column=2).value)
                krets = str(ark.cell(row=rad,column=3).value)

                klubb = datasenter.hentKlubbFraNavn(kjonn,klubbnavn)
                klubb.settKrets(krets)  

                navn = str(ark.cell(row=rad,column=6).value)
                fAar = str(ark.cell(row=rad,column=7).value)

                utover = datasenter.hentUtover(kjonn,navn,fAar)

                poeng = str(ark.cell(row=rad,column=4).value)

                if not poeng.isdigit():
                    raise SystemError(f"Poeng er gitt i heltall ('{poeng}')")

                ovelse = str(ark.cell(row=rad,column=5).value)

                if not ovelse in ovelsesinfo.keys():
                    raise SystemError(f"'{ovelse}' er ikke en gyldig øvelse")

                res = str(ark.cell(row=rad,column=8).value)
                sted = str(ark.cell(row=rad,column=9).value)
                dato = str(ark.cell(row=rad,column=10).value)

                resultat = Resultat(utover,ovelse,poeng,res,dato,sted,klubb)

                utover.leggTilRes(resultat)
                klubb.leggTilRes(resultat)
                datasenter.leggTilRes(kjonn,resultat)